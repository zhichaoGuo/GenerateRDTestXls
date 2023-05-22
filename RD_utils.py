import datetime
import json
from datetime import date
import openpyxl
import requests
import yaml
from openpyxl.styles import Alignment, PatternFill

elpis_repo = ['VOIP',
              'voip/build',
              'voip/customization',
              'voip/manifest',
              'voip/pack',
              'voip/module/webpage',
              'voip/module/vendor',
              'voip/module/minigui',
              'voip/platform/mtk/patch',
              'voip/platform/dspg/kernel',
              'voip/platform/dspg/u-boot',
              'voip/platform/dspg/tools']
hos_repo = ['android/device/htek',
            'external/camera_engine_rkaiq',
            'HtekOEM',
            'platform/frameworks/av',
            'platform/frameworks/base',
            'platform/frameworks/native',
            'platform/packages/apps/Launcher3',
            'platform/packages/apps/Settings',
            'platform/packages/modules/NetworkStack',
            'platform/system/core',
            'platform/system/netd',
            'rk/hardware/rk29/audio',
            'rk/kernel',
            'voip/module/android',
            'vendor/htek/common',
            'vendor/htek/frameworks/apps/VoIP']

with open("cfg.yaml", "r+") as f:
    conf = yaml.safe_load(f)


def generate_time():
    """
    获取当日日期 ->str  2021-11-01
    :return:
    """
    today = date.today()
    return today


def generate_select_time(today):
    """
    根据传入日期推算验单日期（如周一则瑞算至周五）->str 2021-10-29
    :param today:str
    :return select_time:str
    """
    if today.isoweekday() == 1:
        delete_day = datetime.timedelta(days=3)
    else:
        delete_day = datetime.timedelta(days=1)
    select_time = today - delete_day
    return str(select_time)


def count_select_time(start_time, end_time):
    if (start_time is None) & (end_time is None):
        # 不输入起始时间也不输入终止时间，按照今日研发测试进行搜索
        start_time = generate_select_time(generate_time())
        end_time = generate_time()
    elif (start_time is not None) & (end_time is not None):
        end_time = end_time
    else:
        start_time = generate_time()
        end_time = generate_time()
    return start_time, end_time


def get_cookies_without_selenium():
    url = conf['root_url'] + 'login/%2Fq%2Fstatus%3Aopen'
    body = {"username": conf['auth']['username'],
            "password": conf['auth']['password']}
    r = requests.post(url, params=body)
    GerritAccount = r.request.headers.get('Cookie').split('GerritAccount=')[-1]
    XSRF_TOKEN = r.cookies.get('XSRF_TOKEN')
    cookie = 'jenkins-timestamper-offset=-28800000; GerritAccount=' + GerritAccount + '; XSRF_TOKEN=' + XSRF_TOKEN
    cookies = {'Cookie': cookie}
    return cookies


def get_data_from_gerrit(url: str, header: dict) -> list:
    r = requests.get(url, headers=header)
    json_text = r.text
    json_text = json_text[4:-1]
    data_list = json.loads(json_text)
    return data_list


def adjust_repeat_data(data_list: list) -> dict:
    """
    将同一笔patch的不同分支合并成为新的数据结构
    :param data_list:
    :return:
    """
    out = {}
    for i in data_list:
        if i['change_id'] not in out.keys():
            out[i['change_id']] = {'project': i['project'], 'branch': [i['branch']], 'subject': i['subject']}
        else:
            out[i['change_id']]['branch'].append(i['branch'])
    return out


def adjust_original_data(data_list: list) -> (list, list):
    """

    :param data_list:
    :return:
    """
    del_list = ['_number', 'created', 'deletions', 'has_review_started', 'hashtags', 'id', 'insertions',
                'labels', 'owner', 'requirements', 'status', 'submitted', 'submitter', 'total_comment_count',
                'unresolved_comment_count', 'updated']
    _elpis, _hos, _other = [], [], []
    for i in range(len(data_list)):
        for j in del_list:
            del data_list[i][j]
    for i in data_list:
        if i['project'] in elpis_repo:
            _elpis.append(i)
        elif i['project'] in hos_repo:
            _hos.append(i)
        else:
            _other.append(i)
    _elpis = adjust_repeat_data(_elpis)
    _hos = adjust_repeat_data(_hos)
    _other = adjust_repeat_data(_other)
    if _other:
        print('未定义repo库：%s' % _other)
    return _elpis, _hos, _other


def make_excel(elpis: dict, hos: dict, other: dict):
    # 新建工作簿与工作表
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='sheet1')
    sheet1 = wb.worksheets[0]
    work_line_index = 1
    # 设置列宽
    sheet1.column_dimensions['A'].width = 43
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['C'].width = 22
    sheet1.column_dimensions['D'].width = 55
    sheet1.column_dimensions['E'].width = 20

    work_table = []
    for i in [elpis, hos, other]:
        if i:
            work_table.append(i)
    for table in work_table:
        # 设置表头格式
        sheet1.merge_cells(start_row=work_line_index, start_column=1, end_row=work_line_index, end_column=5)
        sheet1.cell(work_line_index, 1).fill = PatternFill("solid", fgColor="6BA9E6")
        title_name = generate_select_time(generate_time()) + ' PATCH 信息'
        sheet1.cell(work_line_index, 1, title_name)
        alignment_center = Alignment(horizontal='center', vertical='center')
        sheet1.cell(work_line_index, 1).alignment = alignment_center
        fill = PatternFill("solid", fgColor="80A2DA")
        fill1 = PatternFill("solid", fgColor="EAF6E1")
        # 设置表头样式
        work_line_index = work_line_index + 1
        sheet1.cell(work_line_index, 1, 'change_id').fill = fill
        sheet1.cell(work_line_index, 2, 'repo').fill = fill
        sheet1.cell(work_line_index, 3, 'branch').fill = fill
        sheet1.cell(work_line_index, 4, 'JIRA ID and DESC').fill = fill
        sheet1.cell(work_line_index, 5, 'Comment').fill = fill
        # 填充数据
        work_line_index = work_line_index + 1
        for i in table:
            for j in table[i]['branch']:
                sheet1.cell(work_line_index, 1, i).fill = fill1
                sheet1.cell(work_line_index, 2, table[i]['project']).fill = fill1
                sheet1.cell(work_line_index, 3, j).alignment = Alignment(wrapText=True)
                sheet1.cell(work_line_index, 3).fill = fill1
                sheet1.cell(work_line_index, 4, table[i]['subject']).alignment = Alignment(wrapText=True)
                sheet1.cell(work_line_index, 4).fill = fill1
                sheet1.cell(work_line_index, 5).fill = fill1
                work_line_index = work_line_index + 1
            # 合并单元格
            if len(table[i]['branch']) > 1:
                sheet1.merge_cells(start_row=work_line_index - len(table[i]['branch']), start_column=1,
                                   end_row=work_line_index - 1, end_column=1)
                sheet1.merge_cells(start_row=work_line_index - len(table[i]['branch']), start_column=2,
                                   end_row=work_line_index - 1, end_column=2)
                sheet1.merge_cells(start_row=work_line_index - len(table[i]['branch']), start_column=4,
                                   end_row=work_line_index - 1, end_column=4)
                sheet1.merge_cells(start_row=work_line_index - len(table[i]['branch']), start_column=5,
                                   end_row=work_line_index - 1, end_column=5)
        # 写入表尾内容并设置格式
        sheet1.cell(work_line_index, 1, '总结').fill = fill
        sheet1.merge_cells(start_row=work_line_index, start_column=2, end_row=work_line_index, end_column=5)
        sheet1.cell(work_line_index, 2, '本次总计测试   个patch 测试出了   个问题，今日合入质量: ').fill = fill
        work_line_index = work_line_index + 2
    # 保存工作表
    wb.save('./ExcelFile/' + str(generate_time()) + '研发测试.xlsx')
